"""format_convert — any-to-any format conversion for xlsx/csv/json."""

from __future__ import annotations

import csv
import os
import tempfile
from pathlib import Path

from tools._path_utils import validate_path
from utils.logging import get_logger

logger = get_logger(__name__)

_SUPPORTED_FORMATS = frozenset({"xlsx", "csv", "json"})


def _sniff_format(path: str) -> str:
    suffix = Path(path).suffix.lower().lstrip(".")
    if suffix in _SUPPORTED_FORMATS:
        return suffix
    return ""


async def format_convert(
    source: str,
    target: str,
    source_fmt: str = "",
    target_fmt: str = "",
) -> dict:
    """Convert a file between xlsx/csv/json formats.

    Args:
        source: Source file path.
        target: Target file path.
        source_fmt: Source format (xlsx/csv/json). Empty = sniff from extension.
        target_fmt: Target format (xlsx/csv/json). Empty = sniff from extension.

    Returns:
        {"ok": True, "result": "已转换: <target>", "target": "..."} or {"ok": False, "error": "..."}
    """
    src_fmt = source_fmt or _sniff_format(source)
    tgt_fmt = target_fmt or _sniff_format(target)

    if src_fmt not in _SUPPORTED_FORMATS:
        return {"ok": False, "error": f"不支持的源格式: {src_fmt}"}
    if tgt_fmt not in _SUPPORTED_FORMATS:
        return {"ok": False, "error": f"不支持的目标格式: {tgt_fmt}"}
    if src_fmt == tgt_fmt:
        return {"ok": False, "error": f"源格式和目标格式相同: {src_fmt}"}

    try:
        src_path = validate_path(source)
        tgt_path = validate_path(target)

        if not src_path.exists():
            return {"ok": False, "error": f"源文件不存在: {source}"}

        if src_fmt == "xlsx" and tgt_fmt == "csv":
            _xlsx_to_csv(src_path, tgt_path)
        elif src_fmt == "csv" and tgt_fmt == "xlsx":
            _csv_to_xlsx(src_path, tgt_path)
        elif src_fmt == "csv" and tgt_fmt == "json":
            await _csv_to_json_delegate(src_path, tgt_path)
        elif src_fmt == "json" and tgt_fmt == "csv":
            await _json_to_csv_delegate(src_path, tgt_path)
        elif src_fmt == "xlsx" and tgt_fmt == "json":
            await _xlsx_to_json_two_step(src_path, tgt_path)
        elif src_fmt == "json" and tgt_fmt == "xlsx":
            await _json_to_xlsx_two_step(src_path, tgt_path)
        else:
            return {"ok": False, "error": f"不支持的转换: {src_fmt} → {tgt_fmt}"}

        return {"ok": True, "result": f"已转换 {source} → {target}（{src_fmt} → {tgt_fmt}）", "target": target}
    except Exception as e:
        logger.warning("format_convert: %s → %s failed: %s", src_fmt, tgt_fmt, e, exc_info=True)
        return {"ok": False, "error": str(e)}


# ── xlsx ↔ csv ──


def _xlsx_to_csv(src: Path, tgt: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(src, read_only=True)
    try:
        ws = wb.active
        tgt.parent.mkdir(parents=True, exist_ok=True)
        fd, tmp_name = tempfile.mkstemp(
            suffix=".csv", prefix=f"_xlsx_to_csv_{tgt.stem}_", dir=str(tgt.parent)
        )
        os.close(fd)
        tmp = Path(tmp_name)
        try:
            with open(tmp, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            os.replace(str(tmp), str(tgt))
        except BaseException:
            if tmp.exists():
                tmp.unlink()
            raise
    finally:
        wb.close()


def _read_csv(src: Path) -> list[list[str]]:
    """Read CSV with automatic encoding detection (utf-8-sig → gbk fallback)."""
    for enc in ("utf-8-sig", "gbk"):
        try:
            with open(src, "r", encoding=enc, newline="") as f:
                return list(csv.reader(f))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法以 utf-8-sig 或 gbk 编码读取 {src}")


def _csv_to_xlsx(src: Path, tgt: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        tgt.parent.mkdir(parents=True, exist_ok=True)
        rows = _read_csv(src)
        for row in rows:
            ws.append(row)
        fd, tmp_name = tempfile.mkstemp(
            suffix=".xlsx", prefix=f"_csv_to_xlsx_{tgt.stem}_", dir=str(tgt.parent)
        )
        os.close(fd)
        tmp = Path(tmp_name)
        try:
            wb.save(str(tmp))
            os.replace(str(tmp), str(tgt))
        except BaseException:
            if tmp.exists():
                tmp.unlink()
            raise
    finally:
        wb.close()


# ── csv ↔ json (delegate to adapters.py) ──


async def _csv_to_json_delegate(src: Path, tgt: Path) -> None:
    from tools.adapters import csv_to_json

    output_dir = str(tgt.parent)
    input_files = {"input": str(src)}
    await csv_to_json(input_files=input_files, output_dir=output_dir)
    # adapters.py generates output based on input filename; rename to target
    generated = tgt.parent / f"{src.stem}.json"
    if generated != tgt and generated.exists():
        import shutil
        shutil.move(str(generated), str(tgt))


async def _json_to_csv_delegate(src: Path, tgt: Path) -> None:
    from tools.adapters import json_to_csv

    output_dir = str(tgt.parent)
    input_files = {"input": str(src)}
    await json_to_csv(input_files=input_files, output_dir=output_dir)
    generated = tgt.parent / f"{src.stem}.csv"
    if generated != tgt and generated.exists():
        import shutil
        shutil.move(str(generated), str(tgt))


# ── xlsx ↔ json (two-step via temp csv) ──


async def _xlsx_to_json_two_step(src: Path, tgt: Path) -> None:
    fd, tmp_name = tempfile.mkstemp(suffix=".csv", prefix=f"_fc_{src.stem}_")
    os.close(fd)
    tmp_csv = Path(tmp_name)
    try:
        _xlsx_to_csv(src, tmp_csv)
        await _csv_to_json_delegate(tmp_csv, tgt)
    finally:
        if tmp_csv.exists():
            tmp_csv.unlink()


async def _json_to_xlsx_two_step(src: Path, tgt: Path) -> None:
    fd, tmp_name = tempfile.mkstemp(suffix=".csv", prefix=f"_fc_{src.stem}_")
    os.close(fd)
    tmp_csv = Path(tmp_name)
    try:
        await _json_to_csv_delegate(src, tmp_csv)
        _csv_to_xlsx(tmp_csv, tgt)
    finally:
        if tmp_csv.exists():
            tmp_csv.unlink()
