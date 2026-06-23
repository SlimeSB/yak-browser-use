"""Tests for format_convert tool."""

import csv
import json
import pytest
from pathlib import Path

from yak_browser_use.tools.format_convert import format_convert


@pytest.mark.asyncio
async def test_xlsx_to_csv(tmp_path):
    import openpyxl

    src = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "age"])
    ws.append(["Alice", "30"])
    wb.save(src)
    wb.close()

    tgt = tmp_path / "data.csv"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    with open(tgt, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["name", "age"]
    assert rows[1] == ["Alice", "30"]


@pytest.mark.asyncio
async def test_csv_to_xlsx(tmp_path):
    import openpyxl

    src = tmp_path / "data.csv"
    src.write_text("name,age\nBob,25\n", encoding="utf-8-sig")

    tgt = tmp_path / "data.xlsx"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    wb = openpyxl.load_workbook(tgt, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("name", "age")
    assert rows[1] == ("Bob", "25")
    wb.close()


@pytest.mark.asyncio
async def test_csv_to_json(tmp_path):
    src = tmp_path / "data.csv"
    src.write_text("name,age\nCarol,28\n", encoding="utf-8-sig")

    tgt = tmp_path / "data.json"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    data = json.loads(tgt.read_text(encoding="utf-8"))
    assert len(data) == 1
    assert data[0]["name"] == "Carol"


@pytest.mark.asyncio
async def test_json_to_csv(tmp_path):
    src = tmp_path / "data.json"
    src.write_text(json.dumps([{"name": "Dave", "age": "35"}]), encoding="utf-8")

    tgt = tmp_path / "data.csv"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    with open(tgt, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["name"] == "Dave"


@pytest.mark.asyncio
async def test_xlsx_to_json_two_step(tmp_path):
    import openpyxl

    src = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["x", "y"])
    ws.append(["1", "2"])
    wb.save(src)
    wb.close()

    tgt = tmp_path / "data.json"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    data = json.loads(tgt.read_text(encoding="utf-8"))
    assert len(data) == 1
    assert data[0]["x"] == "1"


@pytest.mark.asyncio
async def test_json_to_xlsx_two_step(tmp_path):
    import openpyxl

    src = tmp_path / "data.json"
    src.write_text(json.dumps([{"a": "1", "b": "2"}]), encoding="utf-8")

    tgt = tmp_path / "data.xlsx"
    result = await format_convert(str(src), str(tgt))
    assert result["ok"] is True
    assert tgt.exists()
    wb = openpyxl.load_workbook(tgt, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("a", "b")
    assert rows[1] == ("1", "2")
    wb.close()


@pytest.mark.asyncio
async def test_sniff_format():
    result = await format_convert(
        source="data.xlsx", target="out.csv", source_fmt="", target_fmt=""
    )
    assert "源文件不存在" in result.get("error", "")


@pytest.mark.asyncio
async def test_unsupported_format():
    result = await format_convert(
        source="data.png", target="out.csv", source_fmt="png", target_fmt="csv"
    )
    assert result["ok"] is False
    assert "不支持" in result["error"]


@pytest.mark.asyncio
async def test_same_format():
    result = await format_convert(
        source="dummy.csv", target="dummy2.csv", source_fmt="csv", target_fmt="csv"
    )
    assert result["ok"] is False
    assert "相同" in result["error"]
